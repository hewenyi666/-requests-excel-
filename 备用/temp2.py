from openpyxl import load_workbook
wb = load_workbook('./用例.xlsx')
wbs = wb["Sheet1"]
print(wbs['A3'].value)
# list1 = []
# print(wbs.max_row)
# for cell in wbs['A']:
#     list1.append(cell.value)
#     #print(cell.value)
# print(list1)
# for cell2 in wbs['B']:
#     print(cell2.value)
# for cell3 in wbs['C']:
#     print(cell3.value)
#print([i for i in wbs.rows])
# for row in wbs.rows:
#     for cell in row:
#         print(cell.value)
