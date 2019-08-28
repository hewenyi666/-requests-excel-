from openpyxl import load_workbook
wb = load_workbook('./用例.xlsx')
wb.create_sheet("mysheet")
print(wb.sheetnames)

# 选择特定表
# fir = wb.get_sheet_by_name("Sheet2")
fir = wb["Sheet3"]
fir["A2"] = "nihao"
fir["B2"] = "ko"
print(wb.sheetnames)
wb.save("用例.xlsx")
wb.save("用例.xlsx")