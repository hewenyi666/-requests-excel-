from openpyxl import load_workbook
class OperateExcel:
    def __init__(self,path):
        self.excel = load_workbook(path)
        self.mysheet = self.excel.get_sheet_by_name("Sheet1")
    # 获取用例个数
    def get_nums(self):
        num = self.excel.max_row
        return num
    # 获取特定单元格内容
    def get_cell_value(self,row_num,col_num):
        value1 = self.mysheet.cell(row=row_num, column=col_num)
        return value1
if __name__ == "__main__":
    a = OperateExcel('./用例.xlsx',"Sheet1").get_cell_value("A",3)
    print(a)


